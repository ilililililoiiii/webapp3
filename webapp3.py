import base64
import json
import re
import unicodedata
import uuid
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus

import streamlit as st
import streamlit.components.v1 as components


st.set_page_config(
    page_title="전국 특산물 지도",
    page_icon="🗺️",
    layout="wide",
)

BASE_DIR = Path(__file__).resolve().parent
PRODUCT_INFO_PATH = BASE_DIR / "product_info.json"
ERROR_REQUEST_PATH = BASE_DIR / "오류수정요청.xlsx"
RECIPE_BOARD_PATH = BASE_DIR / "recipe_board.json"
RECIPE_IMAGE_DIR = BASE_DIR / "recipe_board_images"
ICON_DIR_CANDIDATES = [BASE_DIR / "icons", BASE_DIR / "아이콘"]
MAP_COMPONENT_DIR = BASE_DIR / "korea_map_component"

REGION_DATA = {
    "서울경기": {
        "color": "#ef9a9a",
        "provinceNames": [
            "서울특별시",
            "경기도",
            "인천광역시"
        ],
        "cities": [
            {
                "name": "양주시",
                "p": "부추"
            },
            {
                "name": "가평시",
                "p": "잣"
            },
            {
                "name": "파주시",
                "p": "장단콩"
            },
            {
                "name": "이천시",
                "p": "이천쌀"
            },
            {
                "name": "안성시",
                "p": "안성배"
            }
        ]
    },
    "강원도": {
        "color": "#a5d6a7",
        "provinceNames": [
            "강원도",
            "강원특별자치도"
        ],
        "cities": [
            {
                "name": "춘천시",
                "p": "닭고기"
            },
            {
                "name": "원주시",
                "p": "복숭아"
            },
            {
                "name": "강릉시",
                "p": "초당콩"
            },
            {
                "name": "동해시",
                "p": "문어"
            },
            {
                "name": "태백시",
                "p": "고랭지배추"
            },
            {
                "name": "속초시",
                "p": "붉은대게"
            },
            {
                "name": "삼척시",
                "p": "장뇌삼,봄굴"
            },
            {
                "name": "평창군",
                "p": "황태,메밀"
            }
        ]
    },
    "충청남도": {
        "color": "#ffe082",
        "provinceNames": [
            "충청남도",
            "대전광역시"
        ],
        "cities": [
            {
                "name": "천안시",
                "p": "호두"
            },
            {
                "name": "공주시",
                "p": "밤"
            },
            {
                "name": "보령시",
                "p": "대하"
            },
            {
                "name": "논산시",
                "p": "딸기"
            },
            {
                "name": "서천시",
                "p": "전어"
            }
        ]
    },
    "충청북도": {
        "color": "#c5e1a5",
        "provinceNames": [
            "충청북도"
        ],
        "cities": [
            {
                "name": "제천시",
                "p": "황기"
            },
            {
                "name": "보은군",
                "p": "대추"
            },
            {
                "name": "영동군",
                "p": "포도"
            },
            {
                "name": "증평군",
                "p": "인삼"
            },
            {
                "name": "괴산군",
                "p": "대학찰옥수수"
            }
        ]
    },
    "전라남도": {
        "color": "#ce93d8",
        "provinceNames": [
            "전라남도",
            "광주광역시"
        ],
        "cities": [
            {
                "name": "목포시",
                "p": "세발낙지,홍어"
            },
            {
                "name": "여수시",
                "p": "돌산갓,멸치"
            },
            {
                "name": "나주시",
                "p": "배"
            },
            {
                "name": "광양시",
                "p": "매실"
            },
            {
                "name": "담양군",
                "p": "죽순"
            },
            {
                "name": "곡성군",
                "p": "토란"
            },
            {
                "name": "구례군",
                "p": "산수유"
            },
            {
                "name": "고흥군",
                "p": "유자"
            },
            {
                "name": "보성군",
                "p": "녹차"
            },
            {
                "name": "해남군",
                "p": "고구마,배추"
            },
            {
                "name": "영암군",
                "p": "무화과"
            },
            {
                "name": "무안군",
                "p": "양파"
            },
            {
                "name": "영광군",
                "p": "굴비"
            },
            {
                "name": "완도군",
                "p": "전복,김"
            },
            {
                "name": "진도군",
                "p": "대파"
            },
            {
                "name": "신안군",
                "p": "천일염"
            }
        ]
    },
    "전라북도": {
        "color": "#ffcc80",
        "provinceNames": [
            "전라북도",
            "전북특별자치도"
        ],
        "cities": [
            {
                "name": "전주시",
                "p": "콩나물"
            },
            {
                "name": "군산시",
                "p": "박대"
            },
            {
                "name": "익산시",
                "p": "서동마"
            },
            {
                "name": "남원시",
                "p": "미꾸라지"
            },
            {
                "name": "완주군",
                "p": "생강"
            },
            {
                "name": "진안군",
                "p": "홍삼대제"
            },
            {
                "name": "무주군",
                "p": "천마"
            },
            {
                "name": "임실군",
                "p": "치즈"
            },
            {
                "name": "순창군",
                "p": "고추장"
            },
            {
                "name": "고창군",
                "p": "복분자,풍천장어"
            },
            {
                "name": "부안군",
                "p": "오디"
            }
        ]
    },
    "경상북도": {
        "color": "#90caf9",
        "provinceNames": [
            "경상북도",
            "대구광역시"
        ],
        "cities": [
            {
                "name": "포항시",
                "p": "과메기"
            },
            {
                "name": "김천시",
                "p": "자두"
            },
            {
                "name": "안동시",
                "p": "소주"
            },
            {
                "name": "상주시",
                "p": "감"
            },
            {
                "name": "문경시",
                "p": "오미자"
            },
            {
                "name": "청송군",
                "p": "사과"
            },
            {
                "name": "의성군",
                "p": "마늘"
            },
            {
                "name": "성주군",
                "p": "참외"
            },
            {
                "name": "봉화군",
                "p": "송이버섯"
            }
        ]
    },
    "경상남도": {
        "color": "#80cbc4",
        "provinceNames": [
            "경상남도",
            "부산광역시",
            "울산광역시"
        ],
        "cities": [
            {
                "name": "창원시",
                "p": "미더덕"
            },
            {
                "name": "통영시",
                "p": "굴"
            },
            {
                "name": "함양군",
                "p": "산양삼"
            },
            {
                "name": "남해군",
                "p": "시금치"
            },
            {
                "name": "함안군",
                "p": "수박"
            },
            {
                "name": "고성군",
                "p": "방울토마토"
            }
        ]
    },
    "제주도": {
        "color": "#b39ddb",
        "provinceNames": [
            "제주특별자치도"
        ],
        "cities": [
            {
                "name": "서귀포시",
                "p": "한라봉,흑돼지"
            }
        ]
    },
    "울릉도": {
        "color": "#b0bec5",
        "provinceNames": [],
        "cities": [
            {
                "name": "울릉군",
                "p": "명이나물,오징어"
            }
        ]
    },
    "독도": {
        "color": "#78909c",
        "provinceNames": [],
        "cities": [
            {
                "name": "독도",
                "p": "독도새우"
            }
        ]
    }
}
REGION_ORDER = [
    "서울경기",
    "강원도",
    "충청남도",
    "충청북도",
    "전라남도",
    "전라북도",
    "경상북도",
    "경상남도",
    "제주도",
    "울릉도",
    "독도"
]
GEO_URLS = [
    "https://raw.githubusercontent.com/southkorea/southkorea-maps/master/kostat/2018/json/skorea-provinces-2018-geo.json",
    "https://raw.githubusercontent.com/southkorea/southkorea-maps/master/kostat/2013/json/skorea_provinces_geo_simple.json"
]

if not MAP_COMPONENT_DIR.exists() or not (MAP_COMPONENT_DIR / "index.html").exists():
    st.error(
        "korea_map_component/index.html 파일을 찾지 못했습니다. "
        "webapp3_custom_click.py와 korea_map_component 폴더를 같은 위치에 두세요."
    )
    st.stop()

korea_map = components.declare_component("korea_map_click_v4", path=str(MAP_COMPONENT_DIR))


def save_error_request(payload: dict) -> tuple[bool, str]:
    """오류수정요청.xlsx에 수정 요청 로그를 시간순으로 누적 저장합니다."""
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        return False, f"openpyxl 라이브러리를 불러오지 못했습니다: {exc}"

    headers = ["접수시간", "권역", "지역", "특산물", "수정내용", "요청ID"]

    try:
        if ERROR_REQUEST_PATH.exists():
            wb = load_workbook(ERROR_REQUEST_PATH)
            ws = wb.active
            current_headers = [cell.value for cell in ws[1][:len(headers)]]
            if ws.max_row < 1 or current_headers != headers:
                ws.insert_rows(1)
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "오류수정요청"
            ws.append(headers)

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(payload.get("region", "")).strip(),
            str(payload.get("city", "")).strip(),
            str(payload.get("product", "")).strip(),
            str(payload.get("requestText", "")).strip(),
            str(payload.get("requestId", "")).strip(),
        ])

        widths = {"A": 21, "B": 14, "C": 16, "D": 18, "E": 46, "F": 26}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        wb.save(ERROR_REQUEST_PATH)
        return True, f"오류수정요청.xlsx 저장 완료: {ERROR_REQUEST_PATH}"
    except Exception as exc:
        return False, f"오류수정요청.xlsx 저장 실패: {exc}"


@st.cache_data(show_spinner=False)
def load_product_info() -> dict:
    if not PRODUCT_INFO_PATH.exists():
        st.warning("product_info.json 파일이 없습니다. 특산물 상세정보 없이 기본 목록만 표시합니다.")
        return {}

    try:
        return json.loads(PRODUCT_INFO_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        st.error(f"product_info.json 형식 오류: {exc}")
        return {}


@st.cache_data(show_spinner=False)
def load_product_icons() -> dict:
    icons = {}
    icon_dir = next((path for path in ICON_DIR_CANDIDATES if path.exists() and path.is_dir()), None)

    if icon_dir is None:
        st.info("icons 또는 아이콘 폴더가 없습니다. 특산물 아이콘 없이 텍스트만 표시합니다.")
        return icons

    for file in icon_dir.rglob("*.png"):
        if file.name.startswith("._"):
            continue
        try:
            key = unicodedata.normalize("NFC", file.stem.strip())
            encoded = base64.b64encode(file.read_bytes()).decode("ascii")
            icons[key] = f"data:image/png;base64,{encoded}"
        except Exception as exc:
            st.warning(f"아이콘 파일을 불러오지 못했습니다: {file.name} / {exc}")

    if "송이버섯" not in icons and "송이" in icons:
        icons["송이버섯"] = icons["송이"]

    return icons


def split_products(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,·]", str(value)) if item.strip()]


def get_product_info(product_info: dict, region: str, city: str, product: str) -> dict:
    return (((product_info or {}).get(region, {}).get(city, {}).get(product, {})) or {})


def product_icon_html(product_icons: dict, product: str) -> str:
    key = unicodedata.normalize("NFC", product.strip())
    src = product_icons.get(key, "")
    if not src:
        return ""
    return f'<img src="{src}" style="width:28px;height:28px;object-fit:contain;vertical-align:middle;margin-right:6px;">'


def render_styles() -> None:
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 1.8rem;
            max-width: 1220px;
        }
        .region-card {
            border: 1px solid #e5e7eb;
            border-radius: 16px;
            padding: 14px 16px;
            background: #ffffff;
            margin-bottom: 12px;
        }
        .product-detail {
            border: 1px solid #e5e7eb;
            border-radius: 10px;
            background: #f9fafb;
            padding: 14px;
            margin-top: 10px;
            margin-bottom: 13px;
        }
        .small-muted {
            color: #6b7280;
            font-size: 13px;
        }
        .recipe-board-wrap {max-width:900px;margin:0 auto;background:#fff;border:1px solid #e5e7eb;border-radius:18px;padding:28px 34px;}
        .recipe-board-head {display:flex;justify-content:space-between;align-items:center;padding-bottom:16px;border-bottom:1px solid #d1d5db;margin-bottom:24px;}
        .recipe-board-title {font-size:24px;font-weight:900;letter-spacing:6px;color:#444;}
        .recipe-sort {font-size:13px;color:#6b7280;}
        .recipe-comment-card {display:flex;gap:14px;margin:24px 0 8px;}
        .recipe-comment-card.reply {margin-left:56px;padding-left:18px;border-left:2px solid #e5e7eb;}
        .recipe-avatar {width:38px;height:38px;border-radius:999px;background:linear-gradient(135deg,#ffe7c2,#f59e0b);flex-shrink:0;}
        .recipe-comment-body {flex:1;}
        .recipe-comment-name {font-size:14px;font-weight:800;color:#374151;}
        .recipe-comment-time {font-size:12px;color:#9ca3af;margin-bottom:9px;}
        .recipe-comment-text {font-size:14px;line-height:1.65;color:#333;white-space:pre-wrap;}
        .recipe-comment-actions {font-size:12px;color:#6b7280;margin-top:8px;display:flex;gap:15px;}

        /* 특산물 카드 하단 버튼 글자 크기/높이 조정 */
        .stLinkButton a,
        .stButton button {
            font-size: 12px !important;
            font-weight: 700 !important;
            min-height: 42px !important;
            padding: 0.25rem 0.35rem !important;
            white-space: nowrap !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_map(selected_region: str | None) -> str | None:
    clicked_region = korea_map(
        region_data=REGION_DATA,
        region_order=REGION_ORDER,
        selected_region=selected_region,
        geo_urls=GEO_URLS,
        default=selected_region,
        key="korea_map_component",
    )
    return clicked_region


def make_widget_key(*parts: object) -> str:
    """Streamlit 위젯 key 중복 방지를 위해 안전한 고유 key를 생성합니다."""
    raw = "_".join(str(part) for part in parts)
    normalized = unicodedata.normalize("NFKC", raw)
    safe = re.sub(r"[^0-9a-zA-Z가-힣_]+", "_", normalized).strip("_")
    return safe or "widget"


def render_product_error_request_popover(region: str, city: str, product: str, key_prefix: str) -> None:
    """각 특산물 카드에 붙는 개별 오류수정요청 입력창입니다."""
    error_popover = st.popover("수정요청", use_container_width=True)

    with error_popover:
        st.subheader("수정요청")
        st.caption(f"{region} · {city} · {product}")

        request_text = st.text_area(
            "수정내용",
            placeholder="예: 소비기한을 냉장 3일로 수정해주세요.",
            key=f"{key_prefix}_request_text",
            height=110,
        )

        if st.button(
            "수정요청 저장",
            type="primary",
            use_container_width=True,
            key=f"{key_prefix}_save_button",
        ):
            if not request_text.strip():
                st.warning("수정내용을 입력해주세요.")
            else:
                payload = {
                    "region": region,
                    "city": city,
                    "product": product,
                    "requestText": request_text.strip(),
                    "requestId": datetime.now().strftime("%Y%m%d%H%M%S%f"),
                }
                ok, msg = save_error_request(payload)
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)



def html_escape_text(value: object) -> str:
    text = str(value or "")
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def load_recipe_posts() -> list[dict]:
    if not RECIPE_BOARD_PATH.exists():
        return []
    try:
        data = json.loads(RECIPE_BOARD_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_recipe_posts(posts: list[dict]) -> None:
    RECIPE_BOARD_PATH.write_text(json.dumps(posts, ensure_ascii=False, indent=2), encoding="utf-8")


def save_recipe_image(uploaded_file):
    if uploaded_file is None:
        return ""
    RECIPE_IMAGE_DIR.mkdir(exist_ok=True)
    suffix = Path(uploaded_file.name).suffix.lower() or ".png"
    filename = f"{uuid.uuid4().hex}{suffix}"
    path = RECIPE_IMAGE_DIR / filename
    path.write_bytes(uploaded_file.getbuffer())
    return str(path.relative_to(BASE_DIR))


def add_recipe_post(author: str, text: str, image_path: str = "", parent_id: str = "", context: dict | None = None) -> None:
    posts = load_recipe_posts()
    posts.append({
        "id": uuid.uuid4().hex,
        "parent_id": parent_id,
        "author": (author or "익명").strip() or "익명",
        "text": (text or "").strip(),
        "image_path": image_path,
        "likes": 0,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "context": context or {},
    })
    save_recipe_posts(posts)




def delete_recipe_post(post_id: str) -> None:
    """댓글을 삭제합니다. 부모 댓글이면 연결된 대댓글도 함께 삭제합니다."""
    posts = load_recipe_posts()
    delete_ids = {post_id}
    changed = True
    while changed:
        changed = False
        for post in posts:
            if post.get("parent_id") in delete_ids and post.get("id") not in delete_ids:
                delete_ids.add(post.get("id"))
                changed = True

    remaining_posts = []
    for post in posts:
        if post.get("id") in delete_ids:
            image_path = post.get("image_path")
            if image_path:
                try:
                    full_path = BASE_DIR / image_path
                    if full_path.exists() and full_path.is_file():
                        full_path.unlink()
                except Exception:
                    pass
        else:
            remaining_posts.append(post)

    save_recipe_posts(remaining_posts)

def open_recipe_board(region: str, city: str, product: str) -> None:
    st.session_state.page = "recipe_board"
    st.session_state.recipe_board_context = {"region": region, "city": city, "product": product}
    st.rerun()


def render_one_recipe_comment(post: dict, is_reply: bool = False) -> None:
    cls = "recipe-comment-card reply" if is_reply else "recipe-comment-card"
    author = html_escape_text(post.get("author", "익명"))
    created_at = html_escape_text(post.get("created_at", ""))
    text = html_escape_text(post.get("text", ""))
    st.markdown(f"""
        <div class="{cls}">
            <div class="recipe-avatar"></div>
            <div class="recipe-comment-body">
                <div class="recipe-comment-name">{author}</div>
                <div class="recipe-comment-time">{created_at}</div>
                <div class="recipe-comment-text">{text}</div>
    """, unsafe_allow_html=True)
    image_path = post.get("image_path")
    if image_path:
        full_path = BASE_DIR / image_path
        if full_path.exists():
            st.image(str(full_path), width=270)
    st.markdown(f'<div class="recipe-comment-actions">❤ {post.get("likes", 0)} Likes <span>↪ Reply</span></div></div></div>', unsafe_allow_html=True)


def render_delete_button(post: dict, is_reply: bool = False) -> None:
    spacer, delete_col = st.columns([0.82, 0.18]) if not is_reply else st.columns([0.76, 0.24])
    with delete_col:
        if st.button("삭제", key=f"delete_recipe_{post.get('id')}", use_container_width=True):
            delete_recipe_post(post.get("id", ""))
            st.rerun()


def render_recipe_board() -> None:
    context = st.session_state.get("recipe_board_context", {})
    context_text = " · ".join([x for x in [context.get("region"), context.get("city"), context.get("product")] if x])

    st.markdown("## 당신의 요리를 자랑해보세요")
    if context_text:
        st.caption(context_text)

    if st.button("← 특산물 지도로 돌아가기", key="back_to_map"):
        st.session_state.page = "map"
        st.rerun()

    posts = load_recipe_posts()
    current_posts = [p for p in posts if p.get("context", {}) == context] if context else posts

    st.markdown('<div class="recipe-board-wrap">', unsafe_allow_html=True)
    st.markdown(f'<div class="recipe-board-head"><div class="recipe-board-title">{len(current_posts)} COMMENTS</div><div class="recipe-sort">Sort by: Newest⌄</div></div>', unsafe_allow_html=True)

    with st.form("new_recipe_comment", clear_on_submit=True):
        author = st.text_input("이름", value="익명")
        text = st.text_area("댓글", placeholder="Write a comment...", height=105)
        image = st.file_uploader("이미지 첨부", type=["png", "jpg", "jpeg", "gif"], key="main_recipe_image")
        submitted = st.form_submit_button("Publish")
        if submitted:
            if text.strip() or image is not None:
                add_recipe_post(author, text, save_recipe_image(image), context=context)
                st.rerun()
            else:
                st.warning("댓글 내용을 입력하거나 이미지를 첨부해주세요.")

    posts = load_recipe_posts()
    current_posts = [p for p in posts if p.get("context", {}) == context] if context else posts
    parents = [p for p in current_posts if not p.get("parent_id")]
    replies: dict[str, list[dict]] = {}
    for post in current_posts:
        if post.get("parent_id"):
            replies.setdefault(post.get("parent_id"), []).append(post)

    for post in reversed(parents):
        render_one_recipe_comment(post)
        render_delete_button(post)
        for reply in replies.get(post["id"], []):
            render_one_recipe_comment(reply, is_reply=True)
            render_delete_button(reply, is_reply=True)
        with st.expander(f"{post.get('author', '익명')}님 댓글에 답글 달기"):
            with st.form(f"reply_form_{post['id']}", clear_on_submit=True):
                r_author = st.text_input("이름", value="익명", key=f"reply_author_{post['id']}")
                r_text = st.text_area("대댓글", placeholder="Replying...", key=f"reply_text_{post['id']}")
                r_image = st.file_uploader("이미지 첨부", type=["png", "jpg", "jpeg", "gif"], key=f"reply_image_{post['id']}")
                if st.form_submit_button("Reply"):
                    if r_text.strip() or r_image is not None:
                        add_recipe_post(r_author, r_text, save_recipe_image(r_image), parent_id=post["id"], context=context)
                        st.rerun()
                    else:
                        st.warning("대댓글 내용을 입력하거나 이미지를 첨부해주세요.")

    st.markdown('</div>', unsafe_allow_html=True)


def render_product_panel(product_info: dict, product_icons: dict, selected_region: str | None) -> None:
    if not selected_region:
        st.markdown(
            '<div class="region-card small-muted">지도에서 권역을 클릭하면 지역 특산물 목록이 표시됩니다.</div>',
            unsafe_allow_html=True,
        )
        return

    region_info = REGION_DATA[selected_region]
    cities = region_info.get("cities", [])
    st.markdown(
        f"""
        <div class="region-card">
            <b>{selected_region}</b>
            <span class="small-muted"> · {len(cities)}개 지역</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    for city_index, city_item in enumerate(cities):
        city = city_item.get("name", "")
        products = split_products(city_item.get("p", ""))

        # 지역명이 사라지지 않도록 Streamlit expander 라벨을 그대로 사용합니다.
        with st.expander(city or "지역명 없음", expanded=False):
            for product_index, product in enumerate(products):
                icon = product_icon_html(product_icons, product)
                info = get_product_info(product_info, selected_region, city, product)
                shelf = info.get("소비기한", "JSON 파일에 소비기한을 입력해주세요.")
                storage = info.get("보관방법", "JSON 파일에 보관방법을 입력해주세요.")
                recipe_link = info.get("레시피링크", "")
                recipe_url = recipe_link or f"https://www.google.com/search?q={quote_plus(product + ' 레시피')}"
                recipe_label = "등록레시피" if recipe_link else "등록레시피"

                st.markdown(
                    f"""
                    <div class="product-detail">
                        <div style="font-size:17px;font-weight:900;margin-bottom:8px;">
                            {icon}{product}
                        </div>
                        <div style="font-size:13px;color:#4b5563;line-height:1.7;">
                            <b>권역</b> {selected_region}<br>
                            <b>지역</b> {city}<br>
                            <b>소비기한</b> {shelf}<br>
                            <b>보관방법</b> {storage}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                key_prefix = make_widget_key(
                    "product_error",
                    selected_region,
                    city,
                    product,
                    city_index,
                    product_index,
                )
                recipe_col, board_col, error_col = st.columns([0.38, 0.36, 0.26], gap="small")
                with recipe_col:
                    st.link_button(
                        recipe_label,
                        recipe_url,
                        use_container_width=True,
                    )
                with board_col:
                    if st.button(
                        "레시피 게시판",
                        use_container_width=True,
                        key=f"{key_prefix}_recipe_board_button",
                    ):
                        open_recipe_board(selected_region, city, product)
                with error_col:
                    render_product_error_request_popover(selected_region, city, product, key_prefix)

def render_error_request_form() -> None:
    bottom_left, bottom_right = st.columns([0.85, 0.15])

    with bottom_right:
        error_popover = st.popover("수정요청", use_container_width=True)

    with error_popover:
        st.subheader("수정요청")
        st.caption("지도에서 확인한 오류를 선택 후 저장하면 오류수정요청.xlsx에 누적됩니다.")

        region_names = list(REGION_DATA.keys())
        default_index = 0
        selected_now = st.session_state.get("selected_region")
        if selected_now in region_names:
            default_index = region_names.index(selected_now)

        selected_region = st.selectbox("권역", region_names, index=default_index, key="error_region")
        city_items = REGION_DATA[selected_region].get("cities", [])
        city_names = [item.get("name", "") for item in city_items]
        selected_city = st.selectbox("지역", city_names, key="error_city")

        selected_city_item = next((item for item in city_items if item.get("name") == selected_city), {})
        product_names = split_products(selected_city_item.get("p", ""))
        selected_product = st.selectbox("특산물", product_names, key="error_product")

        request_text = st.text_area(
            "수정내용",
            placeholder="예: 소비기한을 냉장 3일로 수정해주세요.",
            key="error_request_text",
            height=120,
        )

        if st.button("수정요청 저장", type="primary", use_container_width=True):
            if not request_text.strip():
                st.warning("수정내용을 입력해주세요.")
            else:
                payload = {
                    "region": selected_region,
                    "city": selected_city,
                    "product": selected_product,
                    "requestText": request_text.strip(),
                    "requestId": datetime.now().strftime("%Y%m%d%H%M%S%f"),
                }
                ok, msg = save_error_request(payload)
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)


def main() -> None:
    render_styles()

    if "selected_region" not in st.session_state:
        st.session_state.selected_region = None
    if "page" not in st.session_state:
        st.session_state.page = "map"
    if "recipe_board_context" not in st.session_state:
        st.session_state.recipe_board_context = {}

    if st.session_state.page == "recipe_board":
        render_recipe_board()
        return

    st.title("전국 특산물 지도")
    st.caption("지도 권역을 클릭하면 오른쪽 특산물 목록이 Streamlit 영역에서 갱신됩니다.")

    product_info = load_product_info()
    product_icons = load_product_icons()

    map_col, side_col = st.columns([0.66, 0.34], gap="large")

    with map_col:
        clicked_region = render_map(st.session_state.selected_region)

    if clicked_region and clicked_region != st.session_state.selected_region:
        st.session_state.selected_region = clicked_region
        st.rerun()

    with side_col:
        render_product_panel(product_info, product_icons, st.session_state.selected_region)



if __name__ == "__main__":
    main()
